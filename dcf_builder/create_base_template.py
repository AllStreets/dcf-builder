"""Script to create the base DCF template file."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import config


def create_base_template():
    """Create a pre-formatted base DCF template."""
    wb = Workbook()

    # Remove default sheet
    del wb["Sheet"]

    # Create all sheets with basic formatting
    for sheet_name in config.SHEET_NAMES:
        ws = wb.create_sheet(sheet_name)

        # Set column A width for all sheets
        ws.column_dimensions["A"].width = 25

        # Add sheet title
        ws["A1"] = sheet_name.upper()
        ws["A1"].font = Font(bold=True, size=14)

    # Set Dashboard as active
    wb.active = wb["Dashboard"]

    # Save template
    output_path = config.BASE_TEMPLATE_PATH
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return output_path


if __name__ == "__main__":
    path = create_base_template()
    print(f"Base template created: {path}")
