"""Template generator for creating DCF workbooks."""

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config
from . import data_fetcher as df


# Styles
HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
NUMBER_FORMAT_MILLIONS = '#,##0.0"M"'
NUMBER_FORMAT_PERCENT = "0.0%"
NUMBER_FORMAT_CURRENCY = "$#,##0.00"
NUMBER_FORMAT_MULTIPLE = "0.0x"


class DCFTemplateGenerator:
    """Generates a complete DCF model workbook."""

    def __init__(self, ticker: str):
        self.ticker = ticker.upper()
        self.wb = Workbook()
        self.data = {}
        self.current_year = datetime.now().year

    def fetch_data(self) -> None:
        """Fetch all required data for the model."""
        self.data["info"] = df.get_stock_info(self.ticker)
        self.data["financials"] = df.get_historical_financials(self.ticker)
        self.data["risk_free"] = df.get_risk_free_rate()
        self.data["wacc"] = df.calculate_wacc(self.ticker)

    def generate(self, output_path: Optional[Path] = None) -> Path:
        """Generate the complete DCF workbook."""
        self.fetch_data()

        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Create all sheets
        self._create_dashboard()
        self._create_assumptions()
        self._create_historical()
        self._create_projections()
        self._create_valuation()
        self._create_comps()
        self._create_sensitivity()
        self._create_football_field()

        # Set Dashboard as active
        self.wb.active = self.wb["Dashboard"]

        # Save
        if output_path is None:
            output_path = Path.cwd() / f"DCF_{self.ticker}_{datetime.now():%Y%m%d}.xlsx"

        self.wb.save(output_path)
        return output_path

    def _create_dashboard(self) -> None:
        """Create executive dashboard sheet."""
        ws = self.wb.create_sheet("Dashboard")

        # Title section
        ws.merge_cells("A1:H1")
        ws["A1"] = f"{self.data['info'].get('name', self.ticker)} ({self.ticker})"
        ws["A1"].font = Font(bold=True, size=18)

        ws.merge_cells("A2:H2")
        ws["A2"] = "DCF Valuation Analysis"
        ws["A2"].font = Font(size=14)

        ws["A3"] = f"Generated: {datetime.now():%Y-%m-%d}"
        ws["F3"] = "Scenario:"
        ws["G3"] = "Base Case"
        ws["G3"].fill = INPUT_FILL

        # Key Metrics section
        ws["A5"] = "KEY METRICS"
        ws["A5"].font = HEADER_FONT
        ws["A5"].fill = HEADER_FILL
        ws.merge_cells("A5:C5")

        metrics = [
            ("Current Price", self.data["info"].get("price"), NUMBER_FORMAT_CURRENCY),
            ("Market Cap (M)", self.data["info"].get("market_cap", 0) / 1e6 if self.data["info"].get("market_cap") else None, NUMBER_FORMAT_MILLIONS),
            ("EV/EBITDA", None, NUMBER_FORMAT_MULTIPLE),  # Calculated later
            ("Beta", self.data["info"].get("beta"), "0.00"),
            ("52-Week High", self.data["info"].get("fifty_two_week_high"), NUMBER_FORMAT_CURRENCY),
            ("52-Week Low", self.data["info"].get("fifty_two_week_low"), NUMBER_FORMAT_CURRENCY),
        ]

        for i, (label, value, fmt) in enumerate(metrics, start=6):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"B{i}"].number_format = fmt

        # Valuation Summary section
        ws["E5"] = "VALUATION SUMMARY"
        ws["E5"].font = HEADER_FONT
        ws["E5"].fill = HEADER_FILL
        ws.merge_cells("E5:G5")

        ws["E6"] = "DCF Value per Share"
        ws["F6"] = "=Valuation!E20"  # Link to valuation sheet

        ws["E7"] = "Current Price"
        ws["F7"] = f"=B6"

        ws["E8"] = "Implied Upside"
        ws["F8"] = "=IF(F7>0,(F6-F7)/F7,0)"
        ws["F8"].number_format = NUMBER_FORMAT_PERCENT

        # Column widths
        for col in ["A", "B", "C", "E", "F", "G"]:
            ws.column_dimensions[col].width = 18

    def _create_assumptions(self) -> None:
        """Create assumptions input sheet."""
        ws = self.wb.create_sheet("Assumptions")

        # Header
        ws["A1"] = "DCF MODEL ASSUMPTIONS"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        # Scenario selector
        ws["A3"] = "Scenario"
        ws["B3"] = "Base"
        ws["B3"].fill = INPUT_FILL

        # Company info
        ws["A5"] = "COMPANY INFORMATION"
        ws["A5"].font = HEADER_FONT
        ws["A5"].fill = HEADER_FILL
        ws.merge_cells("A5:D5")

        ws["A6"] = "Ticker"
        ws["B6"] = self.ticker

        ws["A7"] = "Company Name"
        ws["B7"] = self.data["info"].get("name", "")

        # Market data
        ws["A9"] = "MARKET DATA"
        ws["A9"].font = HEADER_FONT
        ws["A9"].fill = HEADER_FILL
        ws.merge_cells("A9:D9")

        market_data = [
            ("Current Price", self.data["info"].get("price")),
            ("Shares Outstanding (M)", (self.data["info"].get("shares_outstanding") or 0) / 1e6),
            ("Market Cap (M)", (self.data["info"].get("market_cap") or 0) / 1e6),
            ("Beta", self.data["info"].get("beta")),
            ("Risk-Free Rate", self.data["risk_free"]),
        ]

        for i, (label, value) in enumerate(market_data, start=10):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"B{i}"].fill = SUBHEADER_FILL

        # WACC Inputs
        ws["A16"] = "WACC INPUTS"
        ws["A16"].font = HEADER_FONT
        ws["A16"].fill = HEADER_FILL
        ws.merge_cells("A16:D16")

        wacc_inputs = [
            ("Risk-Free Rate", self.data["risk_free"], NUMBER_FORMAT_PERCENT),
            ("Equity Risk Premium", config.DEFAULT_EQUITY_RISK_PREMIUM, NUMBER_FORMAT_PERCENT),
            ("Beta", self.data["info"].get("beta"), "0.00"),
            ("Cost of Equity", "=B17+B18*B19", NUMBER_FORMAT_PERCENT),
            ("Cost of Debt (pre-tax)", 0.05, NUMBER_FORMAT_PERCENT),
            ("Tax Rate", config.DEFAULT_TAX_RATE, NUMBER_FORMAT_PERCENT),
            ("Cost of Debt (after-tax)", "=B21*(1-B22)", NUMBER_FORMAT_PERCENT),
            ("Debt/Total Capital", 0.20, NUMBER_FORMAT_PERCENT),
            ("Equity/Total Capital", "=1-B24", NUMBER_FORMAT_PERCENT),
            ("WACC", "=B20*B25+B23*B24", NUMBER_FORMAT_PERCENT),
        ]

        for i, (label, value, fmt) in enumerate(wacc_inputs, start=17):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            if isinstance(value, (int, float)):
                ws[f"B{i}"].fill = INPUT_FILL
            ws[f"B{i}"].number_format = fmt

        # Projection assumptions
        ws["A28"] = "PROJECTION ASSUMPTIONS"
        ws["A28"].font = HEADER_FONT
        ws["A28"].fill = HEADER_FILL
        ws.merge_cells("A28:D28")

        proj_assumptions = [
            ("Projection Years", config.DEFAULT_PROJECTION_YEARS, "0"),
            ("Revenue Growth Rate", 0.05, NUMBER_FORMAT_PERCENT),
            ("EBITDA Margin", 0.20, NUMBER_FORMAT_PERCENT),
            ("D&A % of Revenue", 0.03, NUMBER_FORMAT_PERCENT),
            ("CapEx % of Revenue", 0.04, NUMBER_FORMAT_PERCENT),
            ("NWC % of Revenue", 0.10, NUMBER_FORMAT_PERCENT),
            ("Terminal Growth Rate", config.DEFAULT_TERMINAL_GROWTH, NUMBER_FORMAT_PERCENT),
        ]

        for i, (label, value, fmt) in enumerate(proj_assumptions, start=29):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"B{i}"].fill = INPUT_FILL
            ws[f"B{i}"].number_format = fmt

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15

    def _create_historical(self) -> None:
        """Create historical financials sheet."""
        ws = self.wb.create_sheet("Historical")

        ws["A1"] = "HISTORICAL FINANCIALS"
        ws["A1"].font = Font(bold=True, size=14)

        financials = self.data["financials"]
        years = sorted(financials.get("years", []), reverse=True)[:5]

        if not years:
            ws["A3"] = "No historical data available"
            return

        # Headers
        ws["A3"] = "Income Statement"
        ws["A3"].font = HEADER_FONT
        ws["A3"].fill = HEADER_FILL
        for i, year in enumerate(years):
            col = get_column_letter(i + 2)
            ws[f"{col}3"] = year
            ws[f"{col}3"].font = HEADER_FONT
            ws[f"{col}3"].fill = HEADER_FILL

        # Income statement items
        income_items = [
            ("Revenue", "revenue"),
            ("Gross Profit", "gross_profit"),
            ("EBITDA", "ebitda"),
            ("EBIT", "ebit"),
            ("Net Income", "net_income"),
        ]

        row = 4
        for label, key in income_items:
            ws[f"A{row}"] = label
            for i, year in enumerate(years):
                col = get_column_letter(i + 2)
                value = financials.get("income_statement", {}).get(year, {}).get(key)
                ws[f"{col}{row}"] = value / 1e6 if value else None
                ws[f"{col}{row}"].number_format = NUMBER_FORMAT_MILLIONS
            row += 1

        # Growth rates
        row += 1
        ws[f"A{row}"] = "Growth Rates"
        ws[f"A{row}"].font = HEADER_FONT
        ws[f"A{row}"].fill = SUBHEADER_FILL
        row += 1

        ws[f"A{row}"] = "Revenue Growth"
        for i, year in enumerate(years[:-1]):
            col = get_column_letter(i + 2)
            curr_rev = financials.get("income_statement", {}).get(year, {}).get("revenue")
            prev_rev = financials.get("income_statement", {}).get(years[i + 1], {}).get("revenue")
            if curr_rev and prev_rev and prev_rev != 0:
                ws[f"{col}{row}"] = (curr_rev - prev_rev) / prev_rev
                ws[f"{col}{row}"].number_format = NUMBER_FORMAT_PERCENT
        row += 1

        # Margins
        ws[f"A{row}"] = "EBITDA Margin"
        for i, year in enumerate(years):
            col = get_column_letter(i + 2)
            ebitda = financials.get("income_statement", {}).get(year, {}).get("ebitda")
            revenue = financials.get("income_statement", {}).get(year, {}).get("revenue")
            if ebitda and revenue and revenue != 0:
                ws[f"{col}{row}"] = ebitda / revenue
                ws[f"{col}{row}"].number_format = NUMBER_FORMAT_PERCENT

        # Balance sheet section
        row += 2
        ws[f"A{row}"] = "Balance Sheet"
        ws[f"A{row}"].font = HEADER_FONT
        ws[f"A{row}"].fill = HEADER_FILL
        for i, year in enumerate(years):
            col = get_column_letter(i + 2)
            ws[f"{col}{row}"].font = HEADER_FONT
            ws[f"{col}{row}"].fill = HEADER_FILL
        row += 1

        balance_items = [
            ("Total Assets", "total_assets"),
            ("Total Liabilities", "total_liabilities"),
            ("Total Equity", "total_equity"),
            ("Cash", "cash"),
            ("Total Debt", "total_debt"),
        ]

        for label, key in balance_items:
            ws[f"A{row}"] = label
            for i, year in enumerate(years):
                col = get_column_letter(i + 2)
                value = financials.get("balance_sheet", {}).get(year, {}).get(key)
                ws[f"{col}{row}"] = value / 1e6 if value else None
                ws[f"{col}{row}"].number_format = NUMBER_FORMAT_MILLIONS
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 20
        for i in range(len(years)):
            ws.column_dimensions[get_column_letter(i + 2)].width = 15

    def _create_projections(self) -> None:
        """Create projections sheet."""
        ws = self.wb.create_sheet("Projections")

        ws["A1"] = "FINANCIAL PROJECTIONS"
        ws["A1"].font = Font(bold=True, size=14)

        # Get base year revenue
        financials = self.data["financials"]
        years = sorted(financials.get("years", []), reverse=True)
        base_revenue = 0
        if years:
            base_revenue = financials.get("income_statement", {}).get(years[0], {}).get("revenue", 0) or 0
            base_revenue = base_revenue / 1e6

        proj_years = list(range(self.current_year, self.current_year + 6))

        # Headers
        ws["A3"] = "Projection"
        ws["A3"].font = HEADER_FONT
        ws["A3"].fill = HEADER_FILL
        for i, year in enumerate(proj_years):
            col = get_column_letter(i + 2)
            ws[f"{col}3"] = year
            ws[f"{col}3"].font = HEADER_FONT
            ws[f"{col}3"].fill = HEADER_FILL

        # Projections with formulas
        items = [
            ("Revenue", NUMBER_FORMAT_MILLIONS),
            ("Revenue Growth", NUMBER_FORMAT_PERCENT),
            ("EBITDA", NUMBER_FORMAT_MILLIONS),
            ("EBITDA Margin", NUMBER_FORMAT_PERCENT),
            ("D&A", NUMBER_FORMAT_MILLIONS),
            ("EBIT", NUMBER_FORMAT_MILLIONS),
            ("Less: Taxes", NUMBER_FORMAT_MILLIONS),
            ("NOPAT", NUMBER_FORMAT_MILLIONS),
            ("Plus: D&A", NUMBER_FORMAT_MILLIONS),
            ("Less: CapEx", NUMBER_FORMAT_MILLIONS),
            ("Less: Change in NWC", NUMBER_FORMAT_MILLIONS),
            ("Unlevered FCF", NUMBER_FORMAT_MILLIONS),
        ]

        row = 4
        for label, fmt in items:
            ws[f"A{row}"] = label
            for i, year in enumerate(proj_years):
                col = get_column_letter(i + 2)
                ws[f"{col}{row}"].number_format = fmt

                if label == "Revenue":
                    if i == 0:
                        ws[f"{col}{row}"] = base_revenue
                    else:
                        prev_col = get_column_letter(i + 1)
                        ws[f"{col}{row}"] = f"={prev_col}{row}*(1+Assumptions!$B$30)"
                elif label == "Revenue Growth":
                    ws[f"{col}{row}"] = "=Assumptions!$B$30"
                elif label == "EBITDA":
                    ws[f"{col}{row}"] = f"={col}4*Assumptions!$B$31"
                elif label == "EBITDA Margin":
                    ws[f"{col}{row}"] = "=Assumptions!$B$31"
                elif label == "D&A":
                    ws[f"{col}{row}"] = f"={col}4*Assumptions!$B$32"
                elif label == "EBIT":
                    ws[f"{col}{row}"] = f"={col}6-{col}8"
                elif label == "Less: Taxes":
                    ws[f"{col}{row}"] = f"={col}9*Assumptions!$B$22"
                elif label == "NOPAT":
                    ws[f"{col}{row}"] = f"={col}9-{col}10"
                elif label == "Plus: D&A":
                    ws[f"{col}{row}"] = f"={col}8"
                elif label == "Less: CapEx":
                    ws[f"{col}{row}"] = f"={col}4*Assumptions!$B$33"
                elif label == "Less: Change in NWC":
                    if i == 0:
                        ws[f"{col}{row}"] = 0
                    else:
                        prev_col = get_column_letter(i + 1)
                        ws[f"{col}{row}"] = f"=({col}4-{prev_col}4)*Assumptions!$B$34"
                elif label == "Unlevered FCF":
                    ws[f"{col}{row}"] = f"={col}11+{col}12-{col}13-{col}14"

            row += 1

        # Highlight FCF row
        for i in range(len(proj_years)):
            col = get_column_letter(i + 2)
            ws[f"{col}15"].fill = SUBHEADER_FILL

        # Column widths
        ws.column_dimensions["A"].width = 22
        for i in range(len(proj_years)):
            ws.column_dimensions[get_column_letter(i + 2)].width = 14

    def _create_valuation(self) -> None:
        """Create DCF valuation sheet."""
        ws = self.wb.create_sheet("Valuation")

        ws["A1"] = "DCF VALUATION"
        ws["A1"].font = Font(bold=True, size=14)

        # DCF inputs
        ws["A3"] = "VALUATION INPUTS"
        ws["A3"].font = HEADER_FONT
        ws["A3"].fill = HEADER_FILL
        ws.merge_cells("A3:B3")

        inputs = [
            ("WACC", "=Assumptions!B26", NUMBER_FORMAT_PERCENT),
            ("Terminal Growth Rate", "=Assumptions!B35", NUMBER_FORMAT_PERCENT),
            ("Terminal Year FCF", "=Projections!G15", NUMBER_FORMAT_MILLIONS),
        ]

        row = 4
        for label, formula, fmt in inputs:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = formula
            ws[f"B{row}"].number_format = fmt
            row += 1

        # Terminal value calculation
        row += 1
        ws[f"A{row}"] = "TERMINAL VALUE"
        ws[f"A{row}"].font = HEADER_FONT
        ws[f"A{row}"].fill = HEADER_FILL
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        ws[f"A{row}"] = "Terminal Value"
        ws[f"B{row}"] = "=B6*(1+B5)/(B4-B5)"
        ws[f"B{row}"].number_format = NUMBER_FORMAT_MILLIONS
        row += 1

        # Present value calculations
        row += 1
        ws[f"A{row}"] = "PRESENT VALUE CALCULATION"
        ws[f"A{row}"].font = HEADER_FONT
        ws[f"A{row}"].fill = HEADER_FILL
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        pv_items = [
            ("PV of FCF Year 1", "=Projections!B15/(1+$B$4)^1"),
            ("PV of FCF Year 2", "=Projections!C15/(1+$B$4)^2"),
            ("PV of FCF Year 3", "=Projections!D15/(1+$B$4)^3"),
            ("PV of FCF Year 4", "=Projections!E15/(1+$B$4)^4"),
            ("PV of FCF Year 5", "=Projections!F15/(1+$B$4)^5"),
            ("PV of Terminal Value", "=B9/(1+$B$4)^5"),
        ]

        for label, formula in pv_items:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = formula
            ws[f"B{row}"].number_format = NUMBER_FORMAT_MILLIONS
            row += 1

        # Enterprise and equity value
        row += 1
        ws[f"A{row}"] = "VALUATION SUMMARY"
        ws[f"A{row}"].font = HEADER_FONT
        ws[f"A{row}"].fill = HEADER_FILL
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        ev_row = row
        ws[f"A{row}"] = "Enterprise Value"
        ws[f"B{row}"] = "=SUM(B12:B17)"
        ws[f"B{row}"].number_format = NUMBER_FORMAT_MILLIONS
        ws[f"B{row}"].fill = SUBHEADER_FILL
        row += 1

        ws[f"A{row}"] = "Less: Debt"
        debt = self._get_latest_debt()
        ws[f"B{row}"] = debt / 1e6 if debt else 0
        ws[f"B{row}"].number_format = NUMBER_FORMAT_MILLIONS
        row += 1

        ws[f"A{row}"] = "Plus: Cash"
        cash = self._get_latest_cash()
        ws[f"B{row}"] = cash / 1e6 if cash else 0
        ws[f"B{row}"].number_format = NUMBER_FORMAT_MILLIONS
        row += 1

        eq_row = row
        ws[f"A{row}"] = "Equity Value"
        ws[f"B{row}"] = f"=B{ev_row}-B{ev_row+1}+B{ev_row+2}"
        ws[f"B{row}"].number_format = NUMBER_FORMAT_MILLIONS
        ws[f"B{row}"].fill = SUBHEADER_FILL
        row += 1

        ws[f"A{row}"] = "Shares Outstanding (M)"
        shares = self.data["info"].get("shares_outstanding")
        ws[f"B{row}"] = shares / 1e6 if shares else 0
        ws[f"B{row}"].number_format = "0.0"
        row += 1

        ws[f"A{row}"] = "DCF Value per Share"
        ws[f"B{row}"] = f"=IF(B{row-1}>0,B{eq_row}/B{row-1},0)"
        ws[f"B{row}"].number_format = NUMBER_FORMAT_CURRENCY
        ws[f"B{row}"].fill = INPUT_FILL
        ws[f"B{row}"].font = Font(bold=True)
        row += 1

        ws[f"A{row}"] = "Current Price"
        ws[f"B{row}"] = self.data["info"].get("price")
        ws[f"B{row}"].number_format = NUMBER_FORMAT_CURRENCY
        row += 1

        ws[f"A{row}"] = "Implied Upside/(Downside)"
        ws[f"B{row}"] = f"=IF(B{row-1}>0,(B{row-2}-B{row-1})/B{row-1},0)"
        ws[f"B{row}"].number_format = NUMBER_FORMAT_PERCENT
        ws[f"B{row}"].font = Font(bold=True)

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18

    def _get_latest_debt(self) -> float:
        """Get most recent total debt."""
        financials = self.data["financials"]
        years = sorted(financials.get("years", []), reverse=True)
        if years:
            return financials.get("balance_sheet", {}).get(years[0], {}).get("total_debt") or 0
        return 0

    def _get_latest_cash(self) -> float:
        """Get most recent cash balance."""
        financials = self.data["financials"]
        years = sorted(financials.get("years", []), reverse=True)
        if years:
            return financials.get("balance_sheet", {}).get(years[0], {}).get("cash") or 0
        return 0

    def _create_comps(self) -> None:
        """Create comparable companies sheet (placeholder for peer input)."""
        ws = self.wb.create_sheet("Comps")

        ws["A1"] = "COMPARABLE COMPANY ANALYSIS"
        ws["A1"].font = Font(bold=True, size=14)

        ws["A3"] = "Enter peer tickers below (up to 10):"

        # Header row
        headers = ["Company", "Ticker", "EV (M)", "Revenue (M)", "EBITDA (M)", "EV/Rev", "EV/EBITDA", "P/E"]
        for i, header in enumerate(headers):
            col = get_column_letter(i + 1)
            ws[f"{col}5"] = header
            ws[f"{col}5"].font = HEADER_FONT
            ws[f"{col}5"].fill = HEADER_FILL

        # Input rows for peers
        for row in range(6, 16):
            ws[f"B{row}"].fill = INPUT_FILL

        # Target company row
        ws["A17"] = self.data["info"].get("name", self.ticker)
        ws["B17"] = self.ticker
        ws["A17"].font = Font(bold=True)
        ws["B17"].font = Font(bold=True)

        # Median row
        ws["A18"] = "Median"
        ws["A18"].font = Font(bold=True)
        ws["A18"].fill = SUBHEADER_FILL

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 10
        for i in range(3, 9):
            ws.column_dimensions[get_column_letter(i)].width = 14

    def _create_sensitivity(self) -> None:
        """Create sensitivity analysis sheet."""
        ws = self.wb.create_sheet("Sensitivity")

        ws["A1"] = "SENSITIVITY ANALYSIS"
        ws["A1"].font = Font(bold=True, size=14)

        # WACC vs Terminal Growth matrix
        ws["A3"] = "DCF Value per Share: WACC vs Terminal Growth"
        ws["A3"].font = Font(bold=True)

        # Terminal growth rates (columns)
        growth_rates = [0.015, 0.02, 0.025, 0.03, 0.035]
        for i, rate in enumerate(growth_rates):
            col = get_column_letter(i + 2)
            ws[f"{col}4"] = rate
            ws[f"{col}4"].number_format = NUMBER_FORMAT_PERCENT
            ws[f"{col}4"].fill = HEADER_FILL
            ws[f"{col}4"].font = HEADER_FONT

        # WACC rates (rows)
        wacc_rates = [0.08, 0.09, 0.10, 0.11, 0.12]
        for i, wacc in enumerate(wacc_rates):
            row = 5 + i
            ws[f"A{row}"] = wacc
            ws[f"A{row}"].number_format = NUMBER_FORMAT_PERCENT
            ws[f"A{row}"].fill = HEADER_FILL
            ws[f"A{row}"].font = HEADER_FONT

            # Placeholder values - in real model these would be formulas
            for j in range(len(growth_rates)):
                col = get_column_letter(j + 2)
                ws[f"{col}{row}"].number_format = NUMBER_FORMAT_CURRENCY
                ws[f"{col}{row}"].fill = SUBHEADER_FILL

        ws["A4"] = "WACC \\ TGR"
        ws["A4"].font = Font(bold=True)

        # Column widths
        ws.column_dimensions["A"].width = 15
        for i in range(len(growth_rates)):
            ws.column_dimensions[get_column_letter(i + 2)].width = 12

    def _create_football_field(self) -> None:
        """Create football field data and chart sheet."""
        ws = self.wb.create_sheet("Football Field")

        ws["A1"] = "FOOTBALL FIELD VALUATION"
        ws["A1"].font = Font(bold=True, size=14)

        # Valuation ranges data
        ws["A3"] = "Valuation Method"
        ws["B3"] = "Low"
        ws["C3"] = "Mid"
        ws["D3"] = "High"

        for col in ["A", "B", "C", "D"]:
            ws[f"{col}3"].font = HEADER_FONT
            ws[f"{col}3"].fill = HEADER_FILL

        methods = [
            ("DCF - Bear Case", "=Valuation!B24*0.85", "=Valuation!B24*0.95", "=Valuation!B24"),
            ("DCF - Base Case", "=Valuation!B24", "=Valuation!B24*1.05", "=Valuation!B24*1.15"),
            ("DCF - Bull Case", "=Valuation!B24*1.1", "=Valuation!B24*1.2", "=Valuation!B24*1.35"),
            ("52-Week Range", self.data["info"].get("fifty_two_week_low"), None, self.data["info"].get("fifty_two_week_high")),
        ]

        for i, (method, low, mid, high) in enumerate(methods, start=4):
            ws[f"A{i}"] = method
            ws[f"B{i}"] = low
            if mid:
                ws[f"C{i}"] = mid
            else:
                ws[f"C{i}"] = f"=(B{i}+D{i})/2"
            ws[f"D{i}"] = high

            for col in ["B", "C", "D"]:
                ws[f"{col}{i}"].number_format = NUMBER_FORMAT_CURRENCY

        # Current price reference line
        ws["A9"] = "Current Price"
        ws["B9"] = self.data["info"].get("price")
        ws["B9"].number_format = NUMBER_FORMAT_CURRENCY
        ws["B9"].font = Font(bold=True)

        # Create bar chart
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Valuation Football Field"
        chart.y_axis.title = "Valuation Method"
        chart.x_axis.title = "Price per Share ($)"

        # Data references
        data = Reference(ws, min_col=2, min_row=3, max_col=4, max_row=7)
        categories = Reference(ws, min_col=1, min_row=4, max_row=7)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4

        # Position chart
        ws.add_chart(chart, "F3")

        # Column widths
        ws.column_dimensions["A"].width = 20
        for col in ["B", "C", "D"]:
            ws.column_dimensions[col].width = 12


def generate_dcf_model(ticker: str, output_path: Optional[Path] = None) -> Path:
    """Generate a complete DCF model for a given ticker."""
    generator = DCFTemplateGenerator(ticker)
    return generator.generate(output_path)
