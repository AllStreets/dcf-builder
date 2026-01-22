"""Tests for template_generator module."""

import pytest
from unittest.mock import patch, MagicMock
import tempfile
from pathlib import Path

from dcf_builder.template_generator import DCFTemplateGenerator, generate_dcf_model


@pytest.fixture
def mock_data():
    """Mock data for template generation."""
    return {
        "info": {
            "name": "Apple Inc.",
            "price": 150.0,
            "market_cap": 2500000000000,
            "beta": 1.2,
            "shares_outstanding": 16000000000,
            "fifty_two_week_high": 180.0,
            "fifty_two_week_low": 120.0,
            "enterprise_value": 2600000000000,
        },
        "financials": {
            "years": [2024, 2023, 2022, 2021, 2020],
            "income_statement": {
                2024: {
                    "revenue": 400000000000,
                    "gross_profit": 180000000000,
                    "ebitda": 130000000000,
                    "ebit": 120000000000,
                    "net_income": 100000000000,
                },
                2023: {
                    "revenue": 380000000000,
                    "gross_profit": 170000000000,
                    "ebitda": 125000000000,
                    "ebit": 115000000000,
                    "net_income": 95000000000,
                },
            },
            "balance_sheet": {
                2024: {
                    "total_assets": 350000000000,
                    "total_liabilities": 280000000000,
                    "total_equity": 70000000000,
                    "cash": 50000000000,
                    "total_debt": 100000000000,
                },
            },
        },
        "risk_free": 0.04,
        "wacc": 0.10,
    }


class TestDCFTemplateGenerator:
    """Tests for DCFTemplateGenerator class."""

    @patch("dcf_builder.template_generator.df")
    def test_generator_creates_all_sheets(self, mock_df, mock_data):
        """Test that generator creates all required sheets."""
        mock_df.get_stock_info.return_value = mock_data["info"]
        mock_df.get_historical_financials.return_value = mock_data["financials"]
        mock_df.get_risk_free_rate.return_value = mock_data["risk_free"]
        mock_df.calculate_wacc.return_value = mock_data["wacc"]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_dcf.xlsx"
            generator = DCFTemplateGenerator("AAPL")
            result_path = generator.generate(output_path)

            assert result_path.exists()

            # Load workbook and check sheets
            from openpyxl import load_workbook

            wb = load_workbook(result_path)

            expected_sheets = [
                "Dashboard",
                "Assumptions",
                "Historical",
                "Projections",
                "Valuation",
                "Comps",
                "Sensitivity",
                "Football Field",
            ]

            for sheet_name in expected_sheets:
                assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"

    @patch("dcf_builder.template_generator.df")
    def test_generator_populates_company_info(self, mock_df, mock_data):
        """Test that company info is populated correctly."""
        mock_df.get_stock_info.return_value = mock_data["info"]
        mock_df.get_historical_financials.return_value = mock_data["financials"]
        mock_df.get_risk_free_rate.return_value = mock_data["risk_free"]
        mock_df.calculate_wacc.return_value = mock_data["wacc"]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_dcf.xlsx"
            generator = DCFTemplateGenerator("AAPL")
            result_path = generator.generate(output_path)

            from openpyxl import load_workbook

            wb = load_workbook(result_path)

            # Check Dashboard has company name
            dashboard = wb["Dashboard"]
            assert "Apple Inc." in str(dashboard["A1"].value)
            assert "AAPL" in str(dashboard["A1"].value)

    @patch("dcf_builder.template_generator.df")
    def test_generate_dcf_model_function(self, mock_df, mock_data):
        """Test the convenience function."""
        mock_df.get_stock_info.return_value = mock_data["info"]
        mock_df.get_historical_financials.return_value = mock_data["financials"]
        mock_df.get_risk_free_rate.return_value = mock_data["risk_free"]
        mock_df.calculate_wacc.return_value = mock_data["wacc"]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_dcf.xlsx"
            result_path = generate_dcf_model("AAPL", output_path)

            assert result_path.exists()
            assert result_path.suffix == ".xlsx"


class TestTemplateFormulas:
    """Tests for formula correctness in templates."""

    @patch("dcf_builder.template_generator.df")
    def test_valuation_sheet_has_formulas(self, mock_df, mock_data):
        """Test that valuation sheet contains proper formulas."""
        mock_df.get_stock_info.return_value = mock_data["info"]
        mock_df.get_historical_financials.return_value = mock_data["financials"]
        mock_df.get_risk_free_rate.return_value = mock_data["risk_free"]
        mock_df.calculate_wacc.return_value = mock_data["wacc"]

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_dcf.xlsx"
            generator = DCFTemplateGenerator("AAPL")
            generator.generate(output_path)

            from openpyxl import load_workbook

            wb = load_workbook(output_path, data_only=False)
            valuation = wb["Valuation"]

            # Check that key cells contain formulas
            # Terminal value should reference WACC and growth rate
            tv_cell = valuation["B9"]
            assert tv_cell.value is not None
